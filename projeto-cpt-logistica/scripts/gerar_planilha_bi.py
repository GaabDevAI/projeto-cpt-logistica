from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.append(str(PROJECT_ROOT))

from src.logistics_bi import consolidate_trips
from src.security import (
    escape_spreadsheet_formula,
    safe_project_path,
    sanitize_output_dataframe,
)


INPUT_PATH = safe_project_path(PROJECT_ROOT, "data/dados_ficticios.csv")
OUTPUT_PATH = safe_project_path(PROJECT_ROOT, "outputs/bi_logistico_ficticio.xlsx")

PRIMARY = "1F6FEB"
DARK = "0F3D5E"
LIGHT = "EAF4FF"
WHITE = "FFFFFF"


def style_header(row) -> None:
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=PRIMARY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color=DARK))


def write_dataframe(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
    df = sanitize_output_dataframe(df)
    for row_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=True),
        start_row,
    ):
        for col_idx, value in enumerate(row, start_col):
            ws.cell(row=row_idx, column=col_idx, value=escape_spreadsheet_formula(value))
        if row_idx == start_row:
            style_header(ws[row_idx])


def autosize(ws) -> None:
    for column_cells in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(
            max(width + 2, 12),
            36,
        )


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    raw_df = pd.read_csv(INPUT_PATH)
    trips_df = consolidate_trips(raw_df)

    wb = Workbook()
    dashboard = wb.active
    dashboard.title = "Dashboard BI"
    agenda = wb.create_sheet("Agenda")
    base = wb.create_sheet("Base Consolidada")
    raw = wb.create_sheet("Dados Brutos")
    history = wb.create_sheet("Historico Importacoes")

    dashboard["A1"] = "BI Logistico de Turnos"
    dashboard["A1"].font = Font(size=18, bold=True, color=WHITE)
    dashboard["A1"].fill = PatternFill("solid", fgColor=PRIMARY)
    dashboard.merge_cells("A1:H1")

    dashboard["A3"] = "Atualizado em"
    dashboard["B3"] = datetime.now()
    dashboard["A4"] = "Base"
    dashboard["B4"] = "Dados ficticios para portfolio"

    kpis = {
        "Viagens": trips_df["Viagem"].nunique(),
        "Veiculos": int(trips_df["Qtd Veiculos"].sum()),
        "Destinos": trips_df["Destino"].nunique(),
        "Pedidos": int(trips_df["Pedidos"].sum()),
        "Peso KG": round(trips_df["Peso KG"].sum(), 2),
    }
    for col_idx, (label, value) in enumerate(kpis.items(), 1):
        dashboard.cell(row=6, column=col_idx, value=label)
        dashboard.cell(row=7, column=col_idx, value=value)
    style_header(dashboard[6])

    shift_summary = (
        trips_df.groupby("Turno")["Viagem"]
        .nunique()
        .reindex(["T1", "T2", "T3"], fill_value=0)
        .reset_index(name="Viagens")
    )
    write_dataframe(dashboard, shift_summary, start_row=10, start_col=1)

    chart = BarChart()
    chart.title = "Viagens por Turno"
    chart.y_axis.title = "Viagens"
    chart.x_axis.title = "Turno"
    data = Reference(dashboard, min_col=2, min_row=10, max_row=13)
    categories = Reference(dashboard, min_col=1, min_row=11, max_row=13)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 7
    chart.width = 14
    dashboard.add_chart(chart, "D10")

    agenda_columns = [
        "Data Saida",
        "Hora Saida",
        "Turno",
        "Janela Turno",
        "Placa",
        "Qtd Veiculos",
        "Motorista",
        "Origem",
        "Destino",
        "Chegada Prevista",
        "Lead Time h",
        "Pedidos",
        "Peso KG",
        "CPT",
        "Status",
        "Viagem",
    ]
    write_dataframe(agenda, trips_df[agenda_columns].sort_values("Saida Prevista"))
    write_dataframe(base, trips_df)
    write_dataframe(raw, raw_df)

    history.append(["Data/Hora Importacao", "Arquivo", "Observacao"])
    style_header(history[1])
    history.append([datetime.now(), INPUT_PATH.name, "Geracao ficticia inicial"])

    for ws in [dashboard, agenda, base, raw, history]:
        autosize(ws)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

    for cell in dashboard["6:7"]:
        for item in cell:
            item.fill = PatternFill("solid", fgColor=LIGHT)

    wb.save(OUTPUT_PATH)
    print(f"Planilha gerada em: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
